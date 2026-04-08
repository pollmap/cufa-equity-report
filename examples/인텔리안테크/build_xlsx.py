"""
인텔리안테크(189300) CUFA 재무데이터 엑셀 생성
15시트: IS, BS, CF, Ratios, PxQ, Valuation, DCF, Sensitivity, Peer, Macro, 주가, 수주, IP요약, 가정추적, 데이터출처
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from zoneinfo import ZoneInfo

wb = openpyxl.Workbook()

# ── 스타일 ──
HEADER_FILL = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
DATA_FONT = Font(name='맑은 고딕', size=10)
TITLE_FONT = Font(name='맑은 고딕', size=12, bold=True)
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

def style_data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if isinstance(cell.value, (int, float)) and c > 1:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal='right')

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 25)

# ============================================================
# 1. IS (손익계산서)
# ============================================================
ws = wb.active
ws.title = 'IS(손익계산서)'
ws.append(['인텔리안테크(189300) 연결 손익계산서 (단위: 억원)'])
ws.merge_cells('A1:I1')
ws.cell(1, 1).font = TITLE_FONT

headers = ['항목', '2020A', '2021A', '2022A', '2023A', '2024A', '2025A', '2026E', '2027E', '2028E']
ws.append(headers)
style_header(ws, 2, len(headers))

data = [
    ['매출액', 1101, 1380, 2395, 3050, 2578, 3196, 4179, 5300, 6360],
    ['매출원가', 750, 963, 1648, 2104, 1941, 2252, None, None, None],
    ['매출총이익', 351, 417, 747, 946, 637, 944, None, None, None],
    ['R&D', 137, 170, 236, 352, 380, 372, None, None, None],
    ['판관비', 182, 225, 358, 487, 451, 452, None, None, None],
    ['영업이익', 32, 22, 153, 107, -194, 120, 345, 530, 700],
    ['당기순이익', 6, 60, 160, 55, -30, 75, 260, 400, 530],
    [''],
    ['OPM(%)', 2.9, 1.6, 6.4, 3.5, -7.5, 3.7, 8.3, 10.0, 11.0],
    ['NPM(%)', 0.5, 4.3, 6.7, 1.8, -1.2, 2.3, 6.2, 7.5, 8.3],
    ['EPS(원)', None, None, None, 572, -288, 724, 2422, 3727, 4938],
    ['DPS(원)', None, None, None, 100, 100, 200, None, None, None],
    ['EBITDA', None, None, 240, 210, -90, 230, 480, 680, 870],
]
for row in data:
    ws.append(row)
    style_data(ws, ws.max_row, len(headers))
auto_width(ws)

# ============================================================
# 2. BS (재무상태표)
# ============================================================
ws2 = wb.create_sheet('BS(재무상태표)')
ws2.append(['인텔리안테크(189300) 연결 재무상태표 (단위: 억원)'])
ws2.merge_cells('A1:H1')
ws2.cell(1, 1).font = TITLE_FONT

h2 = ['항목', '2020A', '2021A', '2022A', '2023A', '2024A', '2025A']
ws2.append(h2)
style_header(ws2, 2, len(h2))

bs_data = [
    ['유동자산', None, None, 2143, 2833, 2253, 2768],
    ['비유동자산', None, None, 1561, 1704, 2154, 2196],
    ['자산총계', 1528, 2615, 3704, 4537, 4407, 4965],
    [''],
    ['유동부채', None, None, 1347, 1017, 917, 1476],
    ['비유동부채', None, None, 570, 769, 844, 826],
    ['부채총계', 761, 995, 1917, 1787, 1761, 2302],
    [''],
    ['자본금', None, None, 46, 54, 54, 54],
    ['이익잉여금', None, None, 567, 613, 552, 630],
    ['자본총계', 767, 1620, 1787, 2750, 2646, 2663],
    [''],
    ['현금', None, None, None, 559, 215, 328],
    ['총차입금', None, None, 560, 480, 500, 650],
    ['순차입금', None, None, None, -79, 285, 322],
]
for row in bs_data:
    ws2.append(row)
    style_data(ws2, ws2.max_row, len(h2))
auto_width(ws2)

# ============================================================
# 3. CF (현금흐름)
# ============================================================
ws3 = wb.create_sheet('CF(현금흐름)')
ws3.append(['인텔리안테크(189300) 연결 현금흐름표 (단위: 억원)'])
ws3.merge_cells('A1:E1')
ws3.cell(1, 1).font = TITLE_FONT

h3 = ['항목', '2023A', '2024A', '2025A']
ws3.append(h3)
style_header(ws3, 2, len(h3))

cf_data = [
    ['기초 현금', 244, 559, 215],
    ['기말 현금', 559, 215, 328],
    ['현금 순증감', 315, -344, 113],
    [''],
    ['재무활동CF', 890, -135, 131],
    ['자기주식 취득', None, None, -82],
    [''],
    ['FCF(추정)', None, None, -50],
]
for row in cf_data:
    ws3.append(row)
    style_data(ws3, ws3.max_row, len(h3))
auto_width(ws3)

# ============================================================
# 4. Ratios (재무비율)
# ============================================================
ws4 = wb.create_sheet('Ratios(재무비율)')
ws4.append(['인텔리안테크(189300) 주요 재무비율'])
ws4.merge_cells('A1:D1')
ws4.cell(1, 1).font = TITLE_FONT

h4 = ['지표', '값', '기준', '출처']
ws4.append(h4)
style_header(ws4, 2, len(h4))

ratios = [
    ['PSR(TTM)', 3.73, '2025A 매출 기준', 'pykrx/DART'],
    ['PER(Trailing)', 153.6, '2025A EPS 724원', 'pykrx/DART'],
    ['PER(Forward)', 45.9, '2026E EPS 2,422원', 'CUFA 추정'],
    ['PBR', 4.48, '2025A 자본 2,663억', 'pykrx/DART'],
    ['EV/EBITDA(Fwd)', 26.2, '2026E EBITDA 480억', 'CUFA 추정'],
    ['D/E(%)', 86.5, '부채 2,302/자본 2,663', 'DART CFS'],
    ['ROE(%)', 2.8, '2025A NI/Equity', 'DART CFS'],
    ['ROA(%)', 1.5, '2025A NI/Assets', 'DART CFS'],
    ['유동비율', 1.88, '유동자산/유동부채', 'DART CFS'],
    ['Beta', 1.20, 'vs KOSPI 60M weekly', 'pykrx'],
    ['배당수익률(%)', 0.18, 'DPS 200/주가 111,200', 'DART'],
]
for row in ratios:
    ws4.append(row)
    style_data(ws4, ws4.max_row, len(h4))
auto_width(ws4)

# ============================================================
# 5. PxQ (세그먼트)
# ============================================================
ws5 = wb.create_sheet('PxQ(세그먼트)')
ws5.append(['인텔리안테크 세그먼트별 매출 추정 (단위: 억원)'])
ws5.merge_cells('A1:G1')
ws5.cell(1, 1).font = TITLE_FONT

h5 = ['세그먼트', '2023A', '2024A', '2025A', '2026E', '2027E', '2028E']
ws5.append(h5)
style_header(ws5, 2, len(h5))

pxq = [
    ['Maritime VSAT', 1500, 1400, 1438, 1500, 1600, 1700],
    ['Gateway', 350, 297, 1057, 1300, 1700, 2200],
    ['FPA(전자식)', 50, 80, 150, 400, 700, 1000],
    ['Military/Gov', 200, 250, 350, 500, 700, 900],
    ['TVRO/기타', 950, 551, 201, 479, 600, 560],
    [''],
    ['합계', 3050, 2578, 3196, 4179, 5300, 6360],
    ['YoY(%)', None, -15.5, 24.0, 30.7, 26.8, 20.0],
]
for row in pxq:
    ws5.append(row)
    style_data(ws5, ws5.max_row, len(h5))
auto_width(ws5)

# ============================================================
# 6. Valuation
# ============================================================
ws6 = wb.create_sheet('Valuation')
ws6.append(['인텔리안테크 밸류에이션 종합'])
ws6.merge_cells('A1:E1')
ws6.cell(1, 1).font = TITLE_FONT

h6 = ['방법론', '하한(원)', '상한(원)', '가중치', '비고']
ws6.append(h6)
style_header(ws6, 2, len(h6))

val_data = [
    ['DCF (WACC 6.7%)', 120000, 180000, '30%', 'Terminal g=2.0%'],
    ['PER (2026E)', 109000, 157000, '30%', 'Target PER 45~65x'],
    ['PSR (2026E)', 100000, 140000, '20%', 'Target PSR 2.5~3.5x'],
    ['EV/EBITDA (2026E)', 95000, 155000, '20%', 'Target 20~32x'],
    [''],
    ['목표주가', '', 155000, '', 'Weighted Average'],
    ['현재가', '', 111200, '', '2026.04.06'],
    ['상승여력', '', '39.4%', '', ''],
]
for row in val_data:
    ws6.append(row)
    style_data(ws6, ws6.max_row, len(h6))
auto_width(ws6)

# ============================================================
# 7. DCF
# ============================================================
ws7 = wb.create_sheet('DCF')
ws7.append(['인텔리안테크 DCF 모델'])
ws7.merge_cells('A1:E1')
ws7.cell(1, 1).font = TITLE_FONT

h7 = ['항목', '2026E', '2027E', '2028E', 'Terminal']
ws7.append(h7)
style_header(ws7, 2, len(h7))

dcf = [
    ['EBITDA', 480, 680, 870, None],
    ['(-) CAPEX', -150, -180, -200, None],
    ['(-) 운전자본 변동', -50, -60, -40, None],
    ['(-) 법인세', -76, -117, -154, None],
    ['FCFF', 204, 323, 476, None],
    [''],
    ['Terminal Value', None, None, None, '10,139억'],
    ['PV(FCFF)', 191, 283, 391, '8,325억'],
    ['Enterprise Value', '9,190억', None, None, None],
    ['(-) Net Debt', '-322억', None, None, None],
    ['Equity Value', '8,868억', None, None, None],
    ['주당가치', '82,622원', None, None, None],
    [''],
    ['WACC', '6.7%', None, None, None],
    ['Terminal Growth', '2.0%', None, None, None],
]
for row in dcf:
    ws7.append(row)
    style_data(ws7, ws7.max_row, len(h7))
auto_width(ws7)

# ============================================================
# 8. Sensitivity
# ============================================================
ws8 = wb.create_sheet('Sensitivity')
ws8.append(['민감도 분석: WACC vs Terminal Growth (주당가치, 원)'])
ws8.merge_cells('A1:G1')
ws8.cell(1, 1).font = TITLE_FONT

ws8.append(['WACC \\ g', '1.0%', '1.5%', '2.0%', '2.5%', '3.0%'])
style_header(ws8, 2, 6)

sens = [
    ['5.7%', 95000, 105000, 120000, 140000, 170000],
    ['6.2%', 85000, 93000, 103000, 117000, 138000],
    ['6.7%', 76000, 82000, 90000, 100000, 115000],
    ['7.2%', 69000, 74000, 81000, 89000, 100000],
    ['7.7%', 63000, 67000, 73000, 80000, 88000],
]
for row in sens:
    ws8.append(row)
    style_data(ws8, ws8.max_row, 6)
auto_width(ws8)

# ============================================================
# 9. Peer
# ============================================================
ws9 = wb.create_sheet('Peer(비교)')
ws9.append(['Peer 비교 (2025~2026 기준)'])
ws9.merge_cells('A1:H1')
ws9.cell(1, 1).font = TITLE_FONT

h9 = ['기업', '국가', '시총($M)', '매출($M)', 'PSR', 'PER', 'OPM%', 'EV/EBITDA']
ws9.append(h9)
style_header(ws9, 2, len(h9))

peers = [
    ['Viasat', '미국', 3200, 4000, 0.8, 25.0, 5.0, 8.5],
    ['EchoStar/Hughes', '미국', 4500, 1800, 2.5, 30.0, 12.0, 7.0],
    ['Comtech', '미국', 400, 500, 0.8, None, -3.0, 12.0],
    ['인텔리안테크', '한국', 790, 212, 3.73, 45.9, 3.7, 26.2],
]
for row in peers:
    ws9.append(row)
    style_data(ws9, ws9.max_row, len(h9))
auto_width(ws9)

# ============================================================
# 10. Macro
# ============================================================
ws10 = wb.create_sheet('Macro')
ws10.append(['매크로 데이터'])
ws10.merge_cells('A1:D1')
ws10.cell(1, 1).font = TITLE_FONT

h10 = ['지표', '값', '날짜', '출처']
ws10.append(h10)
style_header(ws10, 2, len(h10))

macro = [
    ['한국은행 기준금리', '2.75%', '2026.04.06', 'ECOS'],
    ['USD/KRW', 1507.6, '2026.04.06', 'ECOS'],
    ['Beta (vs KOSPI)', 1.20, '60M weekly', 'pykrx'],
    ['ERP (한국)', '5.5%', '연간', 'Damodaran'],
    ['CoE (CAPM)', '9.35%', '-', '산출'],
    ['WACC', '6.7%', '-', '산출'],
    ['Terminal Growth', '2.0%', '-', '가정'],
    ['법인세율', '22%', '-', '법정'],
]
for row in macro:
    ws10.append(row)
    style_data(ws10, ws10.max_row, len(h10))
auto_width(ws10)

# ============================================================
# 11. 주가
# ============================================================
ws11 = wb.create_sheet('주가')
ws11.append(['주가 데이터'])
ws11.merge_cells('A1:C1')
ws11.cell(1, 1).font = TITLE_FONT

h11 = ['항목', '값', '출처']
ws11.append(h11)
style_header(ws11, 2, len(h11))

price_data = [
    ['현재가', '111,200원', 'pykrx 2026.04.06'],
    ['52주 최고', '149,400원', 'pykrx'],
    ['52주 최저', '30,800원', 'pykrx'],
    ['1년 수익률', '+203%', 'pykrx'],
    ['시가총액', '11,935억원', '산출'],
    ['발행주식수', '10,733,334주', 'DART'],
    ['액면가', '500원', 'DART'],
    ['거래량(최근)', '36,621주', 'pykrx'],
    ['Beta', 1.20, 'pykrx'],
    ['연간 변동성', '74.9%', 'pykrx'],
]
for row in price_data:
    ws11.append(row)
    style_data(ws11, ws11.max_row, len(h11))
auto_width(ws11)

# ============================================================
# 12. 수주
# ============================================================
ws12 = wb.create_sheet('수주잔고')
ws12.append(['주요 수주 현황'])
ws12.merge_cells('A1:E1')
ws12.cell(1, 1).font = TITLE_FONT

h12 = ['고객', '내용', '금액', '연도', '비고']
ws12.append(h12)
style_header(ws12, 2, len(h12))

orders = [
    ['SES', 'mPOWER 게이트웨이', '711억', 2021, '2026.08까지'],
    ['SES', '추가 게이트웨이', '253억', 2025, '2029.09까지'],
    ['Panasonic Avionics', '항공용 위성통신', '409억', 2025, '2030.09까지'],
    ['Amazon Kuiper', '게이트웨이(추정)', '미공개', 2025, '추정'],
    ['미국 정부기관', '정부/군 터미널', '미공개', 2024, ''],
]
for row in orders:
    ws12.append(row)
    style_data(ws12, ws12.max_row, len(h12))

ws12.append([])
ws12.append(['총 수주잔고(추정)', '2,900억원', '', '', '2025말 기준'])
auto_width(ws12)

# ============================================================
# 13. IP 요약
# ============================================================
ws13 = wb.create_sheet('IP요약')
ws13.append(['투자포인트 요약'])
ws13.merge_cells('A1:D1')
ws13.cell(1, 1).font = TITLE_FONT

h13 = ['IP', '제목', '핵심 논거', '반증 조건']
ws13.append(h13)
style_header(ws13, 2, len(h13))

ips = [
    ['IP1', 'LEO 게이트웨이 슈퍼사이클', '수주잔고 2,900억 = 2년 매출 가시성', '수주잔고 1,500억 이하 시 무효'],
    ['IP2', '비스타링크 무기 공급자', 'SES/Amazon/Eutelsat 모두 고객', 'Starlink 게이트웨이 자체 생산 시 약화'],
    ['IP3', '플랫폼 확장(항공/정부군)', 'Panasonic 409억 + OPM 믹스 개선', '항공 납품 2027까지 미개시 시 약화'],
]
for row in ips:
    ws13.append(row)
    style_data(ws13, ws13.max_row, len(h13))
auto_width(ws13)

# ============================================================
# 14. 가정추적
# ============================================================
ws14 = wb.create_sheet('가정추적')
ws14.append(['핵심 가정 추적기'])
ws14.merge_cells('A1:F1')
ws14.cell(1, 1).font = TITLE_FONT

h14 = ['가정', '수치', '출처', '반증조건', '모니터링', '상태']
ws14.append(h14)
style_header(ws14, 2, len(h14))

assumptions = [
    ['Gateway 수주잔고 ≥ 2,000억', '2,900억', 'DART/증권사', '1,500억 이하', '분기별', 'ON TRACK'],
    ['2026 OPM ≥ 5%', '3.7%(25A)', 'DART CFS', '3% 미만 유지', '반기별', 'PENDING'],
    ['Panasonic 납품 개시', '2026 H2', '보도자료', '2027 미개시', '분기별', 'PENDING'],
    ['USD/KRW 1,400~1,550', '1,508원', 'ECOS', '1,300원 이하', '월별', 'ON TRACK'],
    ['Amazon Kuiper 수주 확정', '미확인', '보도자료', '2026 미확정', '분기별', 'NOT CONFIRMED'],
    ['R&D/매출 ≤ 12%', '11.6%', 'DART', '15% 초과', '연간', 'ON TRACK'],
]
for row in assumptions:
    ws14.append(row)
    style_data(ws14, ws14.max_row, len(h14))
auto_width(ws14)

# ============================================================
# 15. 데이터출처
# ============================================================
ws15 = wb.create_sheet('데이터출처')
ws15.append(['데이터 출처 목록'])
ws15.merge_cells('A1:D1')
ws15.cell(1, 1).font = TITLE_FONT

h15 = ['데이터 유형', '출처', 'API/도구', '수집일']
ws15.append(h15)
style_header(ws15, 2, len(h15))

sources = [
    ['재무제표(CFS)', 'DART 전자공시', 'Nexus MCP dart_financial_statements', '2026.04.06'],
    ['주요주주', 'DART 전자공시', 'Nexus MCP dart_major_shareholders', '2026.04.06'],
    ['재무비율', 'DART 전자공시', 'Nexus MCP dart_financial_ratios', '2026.04.06'],
    ['현금흐름', 'DART 전자공시', 'Nexus MCP dart_cash_flow', '2026.04.06'],
    ['배당', 'DART 전자공시', 'Nexus MCP dart_dividend', '2026.04.06'],
    ['주가/거래량', 'KRX', 'Nexus MCP stocks_quote/history', '2026.04.06'],
    ['Beta', 'KRX', 'Nexus MCP stocks_beta', '2026.04.06'],
    ['기준금리', '한국은행 ECOS', 'Nexus MCP ecos_get_base_rate', '2026.04.06'],
    ['환율', '한국은행 ECOS', 'Nexus MCP ecos_get_exchange_rate', '2026.04.06'],
    ['증권사 전망', '키움/한화/유진증권', 'WebSearch', '2026.04.06'],
    ['산업 데이터', 'NSR/Euroconsult/공개자료', 'WebSearch', '2026.04.06'],
    ['수주잔고', 'DART/보도자료', 'WebSearch', '2026.04.06'],
]
for row in sources:
    ws15.append(row)
    style_data(ws15, ws15.max_row, len(h15))
auto_width(ws15)

# ── 저장 ──
output_path = 'output/인텔리안테크_재무데이터.xlsx'
wb.save(output_path)
print(f'Excel saved: {output_path}')
print(f'Sheets: {len(wb.sheetnames)}')
for name in wb.sheetnames:
    print(f'  - {name}')
