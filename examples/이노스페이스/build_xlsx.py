"""
이노스페이스(462350) CUFA 재무데이터 엑셀 생성
15시트: IS, BS, CF, Ratios, PxQ, Valuation, DCF, Sensitivity, Peer, Macro, 주가, 수주, IP요약, 가정추적, 데이터출처
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
DATA_FONT = Font(name='맑은 고딕', size=10)
TITLE_FONT = Font(name='맑은 고딕', size=12, bold=True)
NUM_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, Alignment(horizontal='center'), THIN_BORDER

def style_data(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.border = DATA_FONT, THIN_BORDER
        if isinstance(cell.value, (int, float)) and c > 1:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal='right')

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 25)

# 1. IS
ws = wb.active
ws.title = 'IS(손익계산서)'
ws.append(['이노스페이스(462350) 연결 손익계산서 (단위: 억원)'])
ws.merge_cells('A1:H1'); ws.cell(1,1).font = TITLE_FONT
h = ['항목', '2022A', '2023A', '2024A', '2025A', '2026E', '2027E', '2028E']
ws.append(h); style_header(ws, 2, len(h))
for row in [
    ['매출액', 3.4, 2.3, 0.15, 27.5, 400, 900, 1600],
    ['매출원가', 1.9, 1.5, 54.2, 177.6, None, None, None],
    ['매출총이익', 1.5, 0.8, -54.0, -150.1, None, None, None],
    ['R&D', 220.4, 110.8, 171.8, 404.9, None, None, None],
    ['판관비', 21.0, 19.6, 39.3, 73.6, None, None, None],
    ['영업이익', -258.2, -159.3, -329.0, -722.5, None, None, None],
    ['당기순이익', -483.4, -832.5, -333.5, -751.0, None, None, None],
    [''],
    ['EPS(원)', -14484, -16450, -1535, None, None, None, None],
    ['EBITDA', -468.3, -811.1, -299.0, -678.2, None, None, None],
]:
    ws.append(row); style_data(ws, ws.max_row, len(h))
auto_width(ws)

# 2. BS
ws2 = wb.create_sheet('BS(재무상태표)')
ws2.append(['이노스페이스(462350) 연결 재무상태표 (단위: 억원)'])
ws2.merge_cells('A1:F1'); ws2.cell(1,1).font = TITLE_FONT
h2 = ['항목', '2022A', '2023A', '2024A', '2025A']
ws2.append(h2); style_header(ws2, 2, len(h2))
for row in [
    ['유동자산', 130, 137, 317, 264],
    ['비유동자산', 105, 103, 242, 437],
    ['자산총계', 235, 240, 559, 701],
    [''], ['유동부채', 621, 23, 81, 103],
    ['비유동부채', 452, 19, 36, 77],
    ['부채총계', 1073, 42, 117, 180],
    [''], ['자본총계', -838, 198, 442, 520],
    [''], ['현금', 118, 53, 209, 217],
    ['총차입금', 15, 14, 28, 49],
]:
    ws2.append(row); style_data(ws2, ws2.max_row, len(h2))
auto_width(ws2)

# 3. CF
ws3 = wb.create_sheet('CF(현금흐름)')
ws3.append(['이노스페이스(462350) 연결 현금흐름표 (단위: 억원)'])
ws3.merge_cells('A1:C1'); ws3.cell(1,1).font = TITLE_FONT
h3 = ['항목', '2024A', '2025A']
ws3.append(h3); style_header(ws3, 2, len(h3))
for row in [['FCF', None, -543]]:
    ws3.append(row); style_data(ws3, ws3.max_row, len(h3))
auto_width(ws3)

# 4. Ratios
ws4 = wb.create_sheet('Ratios(재무비율)')
ws4.append(['이노스페이스(462350) 주요 재무비율'])
ws4.merge_cells('A1:D1'); ws4.cell(1,1).font = TITLE_FONT
h4 = ['지표', '값', '기준', '출처']
ws4.append(h4); style_header(ws4, 2, len(h4))
for row in [
    ['PSR(TTM)', 84.6, '2025A 매출 기준', 'yfinance'],
    ['PER(Trailing)', None, '적자', '-'],
    ['PBR', 4.47, '2025A 자본 520억', 'yfinance/DART'],
    ['D/E(%)', 9.5, '부채/자본', 'DART'],
    ['ROE(%)', -156.1, '2025A', 'yfinance'],
    ['유동비율', 2.56, '유동자산/유동부채', 'DART'],
]:
    ws4.append(row); style_data(ws4, ws4.max_row, len(h4))
auto_width(ws4)

# 5. PxQ
ws5 = wb.create_sheet('PxQ(세그먼트)')
ws5.append(['이노스페이스 사업 세그먼트'])
ws5.merge_cells('A1:C1'); ws5.cell(1,1).font = TITLE_FONT
h5 = ['세그먼트', '설명', '매출 기여']
ws5.append(h5); style_header(ws5, 2, len(h5))
for row in [
    ['위성발사서비스', 'LEO 소형위성 발사', '핵심 (발사 성공 전제)'],
    ['방산 계약', 'LIG넥스원 등 모의발사체', '89억(3년)'],
    ['엔진 기술이전', '하이브리드 엔진 라이선스', '잠재'],
]:
    ws5.append(row); style_data(ws5, ws5.max_row, len(h5))
auto_width(ws5)

# 6. Valuation
ws6 = wb.create_sheet('Valuation')
ws6.append(['이노스페이스 밸류에이션'])
ws6.merge_cells('A1:D1'); ws6.cell(1,1).font = TITLE_FONT
h6 = ['방법', '적용값', '비고', '투자의견']
ws6.append(h6); style_header(ws6, 2, len(h6))
for row in [
    ['PSR', '84.6x', 'Pre-revenue 기업', 'HOLD'],
    ['EV/Revenue', 'N/M', '적자 기업', '-'],
    ['목표주가', 'N/R', '발사 성공 전까지 산정 불가', '-'],
]:
    ws6.append(row); style_data(ws6, ws6.max_row, len(h6))
auto_width(ws6)

# 7. DCF (적자 기업이라 간략)
ws7 = wb.create_sheet('DCF')
ws7.append(['DCF 적용 불가 — Pre-revenue 우주발사 기업'])
ws7.merge_cells('A1:C1'); ws7.cell(1,1).font = TITLE_FONT
ws7.append(['비고', '발사 성공 후 매출 가시성 확보 시 DCF 적용 가능', ''])
auto_width(ws7)

# 8. Sensitivity
ws8 = wb.create_sheet('Sensitivity')
ws8.append(['시나리오 민감도: 발사 성공 횟수 vs 주가'])
ws8.merge_cells('A1:E1'); ws8.cell(1,1).font = TITLE_FONT
h8 = ['발사\\연도', '2026 0회', '2026 2회', '2026 4회', '2026 6회']
ws8.append(h8); style_header(ws8, 2, len(h8))
for row in [
    ['시총(억)', 500, 3000, 5000, 8000],
    ['주가(원)', 3200, 19200, 32000, 51200],
]:
    ws8.append(row); style_data(ws8, ws8.max_row, len(h8))
auto_width(ws8)

# 9. Peer
ws9 = wb.create_sheet('Peer(비교)')
ws9.append(['Peer 비교'])
ws9.merge_cells('A1:G1'); ws9.cell(1,1).font = TITLE_FONT
h9 = ['기업', '국가', '시총($M)', '매출($M)', 'PSR', '발사횟수', '성공률']
ws9.append(h9); style_header(ws9, 2, len(h9))
for row in [
    ['Rocket Lab', '미국/뉴질랜드', 40800, 602, 58, 84, '89.3%'],
    ['Firefly', '미국', 4600, 160, 29, 7, '57%'],
    ['이노스페이스', '한국', 170, 2, 85, 1, '0%'],
]:
    ws9.append(row); style_data(ws9, ws9.max_row, len(h9))
auto_width(ws9)

# 10. Macro
ws10 = wb.create_sheet('Macro')
ws10.append(['매크로 데이터'])
ws10.merge_cells('A1:D1'); ws10.cell(1,1).font = TITLE_FONT
h10 = ['지표', '값', '날짜', '출처']
ws10.append(h10); style_header(ws10, 2, len(h10))
for row in [
    ['한국 우주예산(2026)', '1.12조원', '2026', '과기부'],
    ['LEO 소형위성(2030)', '30,000기', '-', 'Euroconsult'],
    ['SLV 시장(2030)', '$3.6B', '-', 'Stratistics MRC'],
    ['기준금리', '2.75%', '2026.04.06', 'ECOS'],
    ['USD/KRW', '1,507.6', '2026.04.06', 'ECOS'],
]:
    ws10.append(row); style_data(ws10, ws10.max_row, len(h10))
auto_width(ws10)

# 11. 주가
ws11 = wb.create_sheet('주가')
ws11.append(['주가 데이터'])
ws11.merge_cells('A1:C1'); ws11.cell(1,1).font = TITLE_FONT
h11 = ['항목', '값', '출처']
ws11.append(h11); style_header(ws11, 2, len(h11))
for row in [
    ['현재가', '14,940원', 'pykrx 2026.04.06'],
    ['52주 최고', '24,350원', 'pykrx'],
    ['52주 최저', '7,041원', 'pykrx'],
    ['시가총액', '2,325억원', '산출'],
    ['발행주식수', '15,633,862주', 'DART'],
]:
    ws11.append(row); style_data(ws11, ws11.max_row, len(h11))
auto_width(ws11)

# 12. 수주
ws12 = wb.create_sheet('수주잔고')
ws12.append(['주요 수주 현황'])
ws12.merge_cells('A1:E1'); ws12.cell(1,1).font = TITLE_FONT
h12 = ['고객', '내용', '금액', '연도', '비고']
ws12.append(h12); style_header(ws12, 2, len(h12))
for row in [
    ['CShark+EOS Orbit', '3년 위성 36기 발사', '미공개', 2024, '이탈리아/태국'],
    ['MBS(독일)', '위성 LEO 투입 2회+독점유통', '80억', 2024, '2026-2028'],
    ['LIG넥스원', '모의발사체 등 3종', '89억', 2025, '3년 계약'],
    ['미상', '위성발사서비스', '80억', 2025, ''],
]:
    ws12.append(row); style_data(ws12, ws12.max_row, len(h12))
ws12.append([]); ws12.append(['총 수주잔고', '325억원(15건)', '', '', '2025 기준'])
auto_width(ws12)

# 13. IP요약
ws13 = wb.create_sheet('IP요약')
ws13.append(['투자포인트 요약'])
ws13.merge_cells('A1:D1'); ws13.cell(1,1).font = TITLE_FONT
h13 = ['IP', '제목', '핵심', '반증조건']
ws13.append(h13); style_header(ws13, 2, len(h13))
for row in [
    ['IP1', '로켓은 실패하며 배운다', '25t급 엔진 30초 검증, Rocket Lab도 3회 실패', '2026 재발사 실패 시'],
    ['IP2', '3대륙 발사 인프라 선점', '브라질+포르투갈+캐나다 3거점', '수주잔고 200억 이하'],
    ['IP3', '한국 유일의 민간 로켓', '우주예산 1.12조, 방산 89억', '정부 정책 변경 시'],
]:
    ws13.append(row); style_data(ws13, ws13.max_row, len(h13))
auto_width(ws13)

# 14. 가정추적
ws14 = wb.create_sheet('가정추적')
ws14.append(['핵심 가정 추적기'])
ws14.merge_cells('A1:F1'); ws14.cell(1,1).font = TITLE_FONT
h14 = ['가정', '수치', '출처', '반증조건', '모니터링', '상태']
ws14.append(h14); style_header(ws14, 2, len(h14))
for row in [
    ['2026 발사 성공 1회+', '0회(준비중)', 'IR', '0회면 Kill', '월별', 'PENDING'],
    ['현금잔고 50억+', '217억+유증825억', 'DART', '50억 이하', '분기별', 'ON TRACK'],
    ['수주잔고 200억+', '325억', 'DART', '200억 이하', '반기별', 'ON TRACK'],
]:
    ws14.append(row); style_data(ws14, ws14.max_row, len(h14))
auto_width(ws14)

# 15. 데이터출처
ws15 = wb.create_sheet('데이터출처')
ws15.append(['데이터 출처 목록'])
ws15.merge_cells('A1:D1'); ws15.cell(1,1).font = TITLE_FONT
h15 = ['데이터', '출처', 'API/도구', '수집일']
ws15.append(h15); style_header(ws15, 2, len(h15))
for row in [
    ['재무제표(CFS)', 'DART', 'yfinance + DART API', '2026.04.04'],
    ['주가', 'KRX', 'pykrx', '2026.04.06'],
    ['기준금리/환율', 'ECOS', 'Nexus MCP', '2026.04.06'],
    ['산업데이터', 'Euroconsult/NSR', 'WebSearch', '2026.04.04'],
    ['증권사 전망', '메리츠/유진증권', 'WebSearch', '2026.04.04'],
]:
    ws15.append(row); style_data(ws15, ws15.max_row, len(h15))
auto_width(ws15)

output_path = 'output/이노스페이스_재무데이터.xlsx'
wb.save(output_path)
print(f'Excel saved: {output_path} ({len(wb.sheetnames)} sheets)')
