# -*- coding: utf-8 -*-
"""
삼성전자(005930) CUFA 재무데이터 엑셀 생성
10시트: IS, BS, CF, 주당지표, 밸류에이션, Peer비교, 사업부별, DCF민감도, 매출추정, 데이터출처
출처: DART CFS(연결), FnGuide, pykrx, ECOS
"""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 스타일 상수
# ============================================================
HEADER_FILL  = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
HEADER_FONT  = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
DATA_FONT    = Font(name='맑은 고딕', size=10)
TITLE_FONT   = Font(name='맑은 고딕', size=12, bold=True)
RED_FONT     = Font(name='맑은 고딕', size=10, color='CC0000')
EST_FILL     = PatternFill(start_color='F5F0E8', end_color='F5F0E8', fill_type='solid')  # 추정치 배경
THIN_BORDER  = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
)
NUM_FMT      = '#,##0'
PCT_FMT      = '0.0'

EST_COLS = {'2026E', '2027E', '2028E'}  # 추정치 열 구분용


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border    = THIN_BORDER


def style_data(ws, row, cols, est_start_col=None):
    """데이터 행 스타일. est_start_col 이후는 추정치 배경 적용."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        if cell.value is None or cell.value == '':
            cell.font = DATA_FONT
            continue
        is_est = est_start_col is not None and c >= est_start_col
        cell.font = DATA_FONT
        if is_est:
            cell.fill = EST_FILL
        if isinstance(cell.value, (int, float)) and c > 1:
            if cell.value < 0:
                cell.font = Font(name='맑은 고딕', size=10, color='CC0000',
                                 italic=is_est)
                cell.number_format = '#,##0;[Red](#,##0)'
            else:
                cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal='right')


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)


def add_title(ws, title, merge_end_col):
    ws.append([title])
    ws.merge_cells(f'A1:{get_column_letter(merge_end_col)}1')
    ws.cell(1, 1).font = TITLE_FONT


# ============================================================
# 워크북 생성
# ============================================================
wb = openpyxl.Workbook()

# ============================================================
# 시트 1: IS (포괄손익계산서)
# ============================================================
ws1 = wb.active
ws1.title = 'IS(손익계산서)'
add_title(ws1, '삼성전자(005930) 연결 포괄손익계산서 (단위: 억원, 출처: DART CFS)', 8)
h1 = ['항목', '2022A', '2023A', '2024A', '2025A', '2026E', '2027E', '2028E']
ws1.append(h1)
style_header(ws1, 2, len(h1))

is_rows = [
    ['매출액',       3_022_314, 2_589_355, 3_008_709, 3_336_059, 3_607_000, 3_918_000, 4_200_000],
    ['매출원가',     1_900_418, 1_803_886, 1_865_623, 2_022_355, 2_128_000, 2_293_000, 2_436_000],
    ['매출총이익',   1_121_896,   785_469, 1_143_086, 1_313_704, 1_479_000, 1_625_000, 1_764_000],
    ['판매비·관리비',  688_130,   719_799,   815_827,   877_694,   959_000, 1_005_000, 1_044_000],
    ['영업이익',       433_766,    65_670,   327_260,   436_011,   520_000,   620_000,   720_000],
    ['금융수익',       208_290,   161_001,   167_033,   162_403,      None,      None,      None],
    ['금융비용',       190_277,   126_455,   129_857,   117_338,      None,      None,      None],
    ['법인세비용',     -92_136,   -44_808,    30_784,    42_747,      None,      None,      None],
    ['당기순이익',     556_541,   154_871,   344_514,   452_068,   452_000,   525_000,   620_000],
    [],
    ['OPM(%)',             14.4,       2.5,      10.9,      13.1,      14.4,      15.8,      17.1],
    ['NPM(%)',             18.4,       6.0,      11.5,      13.6,      None,      None,      None],
]

for row_data in is_rows:
    ws1.append(row_data)
    if row_data:
        style_data(ws1, ws1.max_row, len(h1), est_start_col=6)

auto_width(ws1)

# ============================================================
# 시트 2: BS (재무상태표)
# ============================================================
ws2 = wb.create_sheet('BS(재무상태표)')
add_title(ws2, '삼성전자(005930) 연결 재무상태표 (단위: 억원, 출처: DART CFS)', 5)
h2 = ['항목', '2022A', '2023A', '2024A', '2025A']
ws2.append(h2)
style_header(ws2, 2, len(h2))

bs_rows = [
    ['유동자산',   2_184_706, 1_959_366, 2_270_623, 2_476_846],
    ['비유동자산', 2_299_539, 2_599_694, 2_874_697, 3_192_575],
    ['총자산',     4_484_245, 4_559_060, 5_145_319, 5_669_421],
    [],
    ['유동부채',     783_449,   757_195,   933_263, 1_064_113],
    ['비유동부채',   153_301,   165_087,   190_136,   242_104],
    ['총부채',       936_750,   922_282, 1_123_399, 1_306_218],
    [],
    ['자기자본',   3_547_496, 3_636_779, 4_021_921, 4_363_203],
    [],
    ['부채비율(%)',      26.4,      25.4,      27.9,      29.9],
    ['유동비율(%)', round(2_184_706/783_449*100, 1), round(1_959_366/757_195*100, 1),
                   round(2_270_623/933_263*100, 1), round(2_476_846/1_064_113*100, 1)],
]

for row_data in bs_rows:
    ws2.append(row_data)
    if row_data:
        style_data(ws2, ws2.max_row, len(h2))

auto_width(ws2)

# ============================================================
# 시트 3: CF (현금흐름표)
# ============================================================
ws3 = wb.create_sheet('CF(현금흐름)')
add_title(ws3, '삼성전자(005930) 연결 현금흐름표 (단위: 억원, 출처: DART CFS)', 5)
h3 = ['항목', '2022A', '2023A', '2024A', '2025A']
ws3.append(h3)
style_header(ws3, 2, len(h3))

cf_rows = [
    ['영업활동CF',   621_813,  441_374,  729_826,  853_151],
    ['투자활동CF',  -316_028, -169_228, -853_817, -685_122],
    ['재무활동CF',  -193_900,  -85_931,  -77_972, -134_780],
    ['FCF',          305_785,  272_146, -124_000,  168_029],
    [],
    ['CAPEX',        316_028,  169_228,  853_817,  685_122],
    ['영업CF/영업이익(%)',
     round(621_813/433_766*100, 1),
     round(441_374/65_670*100,  1),
     round(729_826/327_260*100, 1),
     round(853_151/436_011*100, 1)],
]

for row_data in cf_rows:
    ws3.append(row_data)
    if row_data:
        style_data(ws3, ws3.max_row, len(h3))

auto_width(ws3)

# ============================================================
# 시트 4: 주당지표
# ============================================================
ws4 = wb.create_sheet('주당지표')
add_title(ws4, '삼성전자(005930) 주당 지표 (단위: 원, 출처: FnGuide)', 8)
h4 = ['항목', '2021A', '2022A', '2023A', '2024A', '2025A', '2026E', '2027E']
ws4.append(h4)
style_header(ws4, 2, len(h4))

per_rows = [
    ['EPS(원)',  6_083,  8_057,  2_131,  4_950,  6_564,  7_400,  8_800],
    ['BPS(원)', 45_700, 50_817, 52_002, 57_930, 63_976, 70_100, 77_500],
    ['DPS(원)',   None,  1_444,  1_444,  1_444,  1_700,  1_800,  2_000],
    ['ROE(%)',    None,   17.1,    4.1,    9.0,   10.9,   11.0,   12.0],
    [],
    ['EPS YoY(%)', None,
     round((8_057/6_083-1)*100, 1),
     round((2_131/8_057-1)*100, 1),
     round((4_950/2_131-1)*100, 1),
     round((6_564/4_950-1)*100, 1),
     round((7_400/6_564-1)*100, 1),
     round((8_800/7_400-1)*100, 1)],
    ['배당성향(%)', None, round(1_444/8_057*100,1), round(1_444/2_131*100,1),
     round(1_444/4_950*100,1), round(1_700/6_564*100,1), round(1_800/7_400*100,1),
     round(2_000/8_800*100,1)],
]

for row_data in per_rows:
    ws4.append(row_data)
    if row_data:
        style_data(ws4, ws4.max_row, len(h4), est_start_col=7)

auto_width(ws4)

# ============================================================
# 시트 5: 밸류에이션
# ============================================================
ws5 = wb.create_sheet('밸류에이션')
add_title(ws5, '삼성전자(005930) 밸류에이션 요약 (기준일: 2026-04-08)', 4)

# 현재 밸류에이션
ws5.append([])
ws5.append(['[ 주요 밸류에이션 지표 ]'])
ws5.cell(ws5.max_row, 1).font = Font(name='맑은 고딕', size=11, bold=True)
h5a = ['지표', '값', '기준', '비고']
ws5.append(h5a)
style_header(ws5, ws5.max_row, len(h5a))

val_rows = [
    ['현재가(원)',    '210,000', '2026-04-08', 'pykrx KRX'],
    ['시가총액(억원)', '12,536,543', '보통주 기준', '= 1,253.7조원'],
    ['PER(TTM)',      '32.0x', '2025A EPS 6,564원', 'FnGuide'],
    ['PER(Fwd)',      '28.4x', '2026E EPS 7,400원', 'FnGuide 컨센서스'],
    ['PBR',          '3.28x', '2025A BPS 63,976원', 'FnGuide'],
    ['PSR(TTM)',      '3.76x', '2025A 매출 333.6조', '산출'],
    ['EV/EBITDA(Fwd)', '12.1x', '2026E EBITDA 10.0조', 'FnGuide'],
    ['배당수익률',    '0.81%', 'DPS 1,700원 / 현재가', '2025A 기준'],
    ['52주 최고',    '223,000원', '', 'pykrx'],
    ['52주 최저',    '52,900원', '', 'pykrx'],
]

for row_data in val_rows:
    ws5.append(row_data)
    style_data(ws5, ws5.max_row, len(h5a))

# WACC 섹션
ws5.append([])
ws5.append(['[ WACC 상세 ]'])
ws5.cell(ws5.max_row, 1).font = Font(name='맑은 고딕', size=11, bold=True)
h5b = ['파라미터', '값', '산출 방법', '출처']
ws5.append(h5b)
style_header(ws5, ws5.max_row, len(h5b))

wacc_rows = [
    ['무위험수익률(Rf)',  '2.75%', '국고채 3년',                '한국은행 ECOS'],
    ['시장위험프리미엄(ERP)', '5.50%', 'Damodaran 한국 2026', 'damodaran.com'],
    ['베타(β)',          '1.15',  'KRX 60M 주간수익률 회귀', 'KRX'],
    ['자기자본비용(Ke)', '9.07%', 'Ke = 2.75 + 1.15×5.5',    '산출'],
    ['타인자본비용(Kd)', '3.80%', '삼성전자 회사채 YTM',      '추정'],
    ['실효법인세율(t)',   '25.0%', '최근 3개년 평균',           'DART'],
    ['부채비율(D/V)',     '7.0%',  '2025A 기준',               'DART'],
    ['자기자본비율(E/V)', '93.0%', '2025A 기준',               'DART'],
    ['WACC',             '8.64%', '9.07×0.93 + 3.80×0.75×0.07', '산출'],
    ['영구성장률(g)',     '2.00%', '한국 명목GDP 성장률',       '한국은행'],
]

for row_data in wacc_rows:
    ws5.append(row_data)
    style_data(ws5, ws5.max_row, len(h5b))

# 시나리오
ws5.append([])
ws5.append(['[ 목표주가 시나리오 ]'])
ws5.cell(ws5.max_row, 1).font = Font(name='맑은 고딕', size=11, bold=True)
h5c = ['시나리오', '목표주가(원)', '확률', '업사이드(%)', '근거']
ws5.append(h5c)
style_header(ws5, ws5.max_row, len(h5c))

scenario_rows = [
    ['Bull', 280_000, '20%', '+33.3%', 'HBM4 점유율 40% + 파운드리 애플 수주 + Korea Discount 완화'],
    ['Base', 250_000, '60%', '+19.0%', '컨센서스 달성, HBM 점유율 35%, 파운드리 BEP 2027E'],
    ['Bear', 160_000, '20%', '-23.8%', '메모리 공급과잉 재발 + HBM 점유율 20% 이하 + 환율 역풍'],
    ['가중평균 목표주가', 234_000, '100%', '+11.4%', '= 280k×20% + 250k×60% + 160k×20%'],
]

for row_data in scenario_rows:
    ws5.append(row_data)
    style_data(ws5, ws5.max_row, len(h5c))

auto_width(ws5)

# ============================================================
# 시트 6: Peer 비교
# ============================================================
ws6 = wb.create_sheet('Peer비교')
add_title(ws6, 'Peer 비교 (2025A 기준, 출처: FnGuide/Bloomberg)', 8)
h6 = ['기업', '국가', 'PER(Fwd)', 'PBR', 'OPM(%)', 'ROE(%)', '시총(조원)', '비고']
ws6.append(h6)
style_header(ws6, 2, len(h6))

peer_rows = [
    ['삼성전자(005930)', '한국',  28.4, 3.28, 13.1, 10.9, 1253.7, '글로벌 1위 메모리 / 갤럭시 생태계'],
    ['SK하이닉스(000660)', '한국', 12.0, 2.5,  38.0, 35.0,  730.0, 'HBM M/S 1위(70%), 메모리 전문'],
    ['TSMC(TSM)',       '대만',   22.0, 7.2,  42.0, 29.0, 1060.0, '파운드리 글로벌 1위(M/S 60%)'],
    ['Micron(MU)',      '미국',   10.0, 2.0,  30.0, 22.0,  120.0, 'DRAM 3위, HBM 신흥 경쟁자'],
    ['Intel(INTC)',     '미국',   25.0, 1.2,   5.0,  2.0,  100.0, '파운드리 IFS 3위, 구조조정 중'],
]

for row_data in peer_rows:
    ws6.append(row_data)
    style_data(ws6, ws6.max_row, len(h6))

# 평균 행
ws6.append(['Peer 평균(삼성 제외)', '',
            round((12.0+22.0+10.0+25.0)/4, 1),
            round((2.5+7.2+2.0+1.2)/4, 2),
            round((38.0+42.0+30.0+5.0)/4, 1),
            round((35.0+29.0+22.0+2.0)/4, 1),
            '', '단순평균'])
style_data(ws6, ws6.max_row, len(h6))
ws6.cell(ws6.max_row, 1).font = Font(name='맑은 고딕', size=10, bold=True)

auto_width(ws6)

# ============================================================
# 시트 7: 사업부별
# ============================================================
ws7 = wb.create_sheet('사업부별')
add_title(ws7, '삼성전자 사업부별 매출 추정 (단위: 억원, 출처: DART/FnGuide)', 6)
h7 = ['사업부', '2023A', '2024A', '2025A(추정)', '2026E', '2027E']
ws7.append(h7)
style_header(ws7, 2, len(h7))

# 2025A 사업부별 합계 기준 비율 적용 추정
# DS:MX:VD/DA:SDC:Harman = 33:36:17:9:5 (config.py 기준)
# 2025A 연결 매출 3,336,059억
seg_rows = [
    ['DS(반도체)',       905_274, 1_050_048, 1_111_000, 1_246_000, 1_406_000],
    ['MX(스마트폰·웨어러블)', 1_031_856, 1_099_888, 1_203_000, 1_307_000, 1_444_000],
    ['VD/DA(가전·TV)',    449_000,   494_000,   552_000,   594_000,   641_000],
    ['SDC(디스플레이)',    238_000,   270_000,   315_000,   331_000,   368_000],
    ['Harman(오디오·차량)', 136_000,   145_000,   153_000,   163_000,   174_000],
    [],
    ['합계(내부거래 前)', 2_760_130, 3_058_936, 3_334_000, 3_641_000, 4_033_000],
    ['연결 매출(내부거래 後)', 2_589_355, 3_008_709, 3_336_059, 3_607_000, 3_918_000],
]

for row_data in seg_rows:
    ws7.append(row_data)
    if row_data:
        style_data(ws7, ws7.max_row, len(h7), est_start_col=5)

ws7.append([])
ws7.append(['주) 2025A 사업부 수치는 내부 추정값. 확정치는 DART 사업보고서(2026.03 공시) 참조 요망.'])
ws7.cell(ws7.max_row, 1).font = Font(name='맑은 고딕', size=9, color='666666')

auto_width(ws7)

# ============================================================
# 시트 8: DCF 민감도
# ============================================================
ws8 = wb.create_sheet('DCF민감도')
add_title(ws8, 'DCF 민감도 분석 — WACC × 영구성장률(g) 매트릭스 (목표주가, 원)', 8)

# WACC: 7.64% ~ 9.64% (기준 8.64%), g: 1.0% ~ 3.0% (기준 2.0%)
# 간소화된 DDM/DCF: TV = FCFF_2028E / (WACC - g), PV 역산
# FCFF 추정: 2028E 순이익 620,000억 + D&A 420,000억 - CAPEX 700,000억 = 340,000억
# 주식수: 5,969,782,550주 = 59.7억주

BASE_FCFF  = 340_000  # 억원 (2028E)
SHARES     = 59.698   # 억주 (보통주 기준)
NET_DEBT   = -468_000  # 억원 (순현금: 현금 520,000 - 순차입금 52,000)

wacc_list = [7.64, 8.14, 8.64, 9.14, 9.64]
g_list    = [1.00, 1.50, 2.00, 2.50, 3.00]

# 헤더 (g 행)
header_row = ['WACC \\ g(%)'] + [f'{g:.2f}%' for g in g_list]
ws8.append(header_row)
style_header(ws8, ws8.max_row, len(header_row))

for wacc in wacc_list:
    row = [f'{wacc:.2f}%']
    for g in g_list:
        if wacc <= g:
            row.append('N/A')
        else:
            # TV = FCFF / (WACC - g) → EV → Equity → 주당
            tv   = BASE_FCFF / ((wacc - g) / 100)
            # 3년 할인 (2025 → 2028)
            pv   = tv / (1 + wacc / 100) ** 3
            eq   = pv + NET_DEBT
            per_share = round(eq / SHARES)
            row.append(per_share)
    ws8.append(row)
    style_data(ws8, ws8.max_row, len(header_row))

ws8.append([])
ws8.append(['가정: 2028E FCFF = 340,000억원 (순이익 620,000 + D&A 420,000 - CAPEX 700,000)',
            '', '현재가: 210,000원', '기준 WACC: 8.64%', 'g: 2.00%', '', '', ''])
ws8.cell(ws8.max_row, 1).font = Font(name='맑은 고딕', size=9, color='666666')

# 기준값 강조 (WACC=8.64%, g=2.00% → row 4, col 4)
try:
    ref_cell = ws8.cell(row=5, column=4)  # 헤더 row=2 + wacc_idx=3 → row=5
    ref_cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    ref_cell.font = Font(name='맑은 고딕', size=10, bold=True)
except Exception:
    pass

auto_width(ws8)

# ============================================================
# 시트 9: 매출추정 (P×Q Bottom-up)
# ============================================================
ws9 = wb.create_sheet('매출추정')
add_title(ws9, '삼성전자 매출 Bottom-up 추정 가정 (FnGuide 컨센서스 기반)', 5)

ws9.append([])
ws9.append(['[ DS사업부 — 메모리 P×Q 가정 ]'])
ws9.cell(ws9.max_row, 1).font = Font(name='맑은 고딕', size=11, bold=True)
h9a = ['항목', '2025A', '2026E', '2027E', '비고']
ws9.append(h9a)
style_header(ws9, ws9.max_row, len(h9a))

ds_rows = [
    ['DRAM 출하(Gb eq, 십억)',         95,   108,   122, '연간 +14% 가정'],
    ['DRAM ASP($/Gb)',                 3.5,   3.8,   4.2, 'HBM 믹스 개선'],
    ['DRAM 매출(억원)',           792_750, 975_600, 1_215_180, '출하×ASP×환율'],
    ['NAND 출하(Gb eq, 십억)',         400,   460,   520, '연간 +15% 가정'],
    ['NAND ASP($/Gb)',                0.08,  0.09,  0.10, '공급 안정화'],
    ['NAND 매출(억원)',           241_600, 311_220, 391_040, '출하×ASP×환율'],
    ['HBM 매출(억원)',            120_000, 200_000, 280_000, 'HBM3E/4 프리미엄'],
    ['시스템반도체+파운드리(억원)',  130_000, 160_000, 200_000, 'Exynos+외부 수주'],
    ['DS 합계(억원)',           1_284_350, 1_646_820, 2_086_220, '합산'],
]

for row_data in ds_rows:
    ws9.append(row_data)
    style_data(ws9, ws9.max_row, len(h9a), est_start_col=3)

ws9.append([])
ws9.append(['[ MX사업부 — 스마트폰 P×Q 가정 ]'])
ws9.cell(ws9.max_row, 1).font = Font(name='맑은 고딕', size=11, bold=True)
ws9.append(h9a)
style_header(ws9, ws9.max_row, len(h9a))

mx_rows = [
    ['갤럭시 출하(백만대)',            230,   236,   242, '글로벌 M/S 20% 유지'],
    ['플래그십 비중(%)',                28,    30,    32, 'S/Z시리즈 ASP 상승'],
    ['평균 ASP($)',                   340,   360,   380, 'Galaxy AI 프리미엄'],
    ['MX 매출(억원)',            1_070_840, 1_216_800, 1_314_768, 'P×Q 산출'],
    ['웨어러블·태블릿(억원)',          132_160,  90_200,  129_232, '보완재 효과'],
    ['MX 합계(억원)',            1_203_000, 1_307_000, 1_444_000, '추정'],
]

for row_data in mx_rows:
    ws9.append(row_data)
    style_data(ws9, ws9.max_row, len(h9a), est_start_col=3)

ws9.append([])
ws9.append(['[ 핵심 가정 — 환율 ]'])
ws9.cell(ws9.max_row, 1).font = Font(name='맑은 고딕', size=11, bold=True)
ws9.append(h9a)
style_header(ws9, ws9.max_row, len(h9a))

macro_rows = [
    ['USD/KRW(원)',  1_390, 1_380, 1_350, '한국은행 전망 (ECOS)'],
    ['HBM 시장 CAGR', '연간+90%', '→ 2027E', '~$80B', 'Gartner 추정'],
]

for row_data in macro_rows:
    ws9.append(row_data)
    style_data(ws9, ws9.max_row, len(h9a))

auto_width(ws9)

# ============================================================
# 시트 10: 데이터 출처
# ============================================================
ws10 = wb.create_sheet('데이터출처')
add_title(ws10, '삼성전자 재무데이터 출처 목록 (수집일: 2026-04-08)', 5)
h10 = ['데이터 항목', '출처', 'API/도구', '수집일', '비고']
ws10.append(h10)
style_header(ws10, 2, len(h10))

src_rows = [
    ['연결 재무제표(CFS)',      'DART',            'DART 전자공시 Open API',       '2026-04-08', '사업보고서 2025.03 공시'],
    ['손익계산서(IS)',          'DART CFS',        'DART OpenAPI / FnGuide',       '2026-04-08', '연결기준 확인 필수'],
    ['재무상태표(BS)',          'DART CFS',        'DART OpenAPI / FnGuide',       '2026-04-08', '연결기준'],
    ['현금흐름표(CF)',          'DART CFS',        'DART OpenAPI',                 '2026-04-08', '연결기준'],
    ['현재주가',                'KRX',             'pykrx',                        '2026-04-08', '종가 기준'],
    ['52주 고/저',             'KRX',             'pykrx',                        '2026-04-08', ''],
    ['시가총액',                'KRX',             'pykrx / 자체 산출',            '2026-04-08', '보통주 기준'],
    ['EPS/BPS/DPS(E)',         'FnGuide',         'FnGuide 컨센서스',             '2026-04-08', '추정치 변동 가능'],
    ['PER/PBR/EV-EBITDA',     'FnGuide',         'FnGuide / 자체 산출',          '2026-04-08', 'Forward 기준 명시'],
    ['Peer 데이터',            'FnGuide/Bloomberg', 'FnGuide 비교분석',            '2026-04-08', '2025A 기준'],
    ['무위험수익률(Rf)',        '한국은행 ECOS',   'Nexus Finance MCP / ECOS API', '2026-04-08', '국고채 3년'],
    ['시장위험프리미엄(ERP)',   'Damodaran',       'damodaran.com',                '2026.01',    '한국 ERP'],
    ['베타(β)',                 'KRX',             'KRX 60M 주간수익률 회귀',      '2026-04-08', '추정치'],
    ['환율 가정',               '한국은행',        'ECOS API',                     '2026-04-08', 'USD/KRW 전망'],
    ['HBM 시장 전망',          'Gartner',         'WebSearch',                    '2026-04',    '추정 CAGR'],
    ['사업부별 매출',          'DART/FnGuide',    '추정 (공식 세그먼트 미분리)',   '2026-04-08', '내부 추정 — 오차 존재'],
]

for row_data in src_rows:
    ws10.append(row_data)
    style_data(ws10, ws10.max_row, len(h10))

ws10.append([])
ws10.append(['★ 주의사항: 추정치(E)는 FnGuide 컨센서스 기반이며 실제치와 상이할 수 있습니다. '
             '투자 결정 시 반드시 최신 DART 공시 및 증권사 리포트를 함께 참조하십시오.'])
ws10.cell(ws10.max_row, 1).font = Font(name='맑은 고딕', size=9, bold=True, color='CC0000')

auto_width(ws10)

# ============================================================
# 저장
# ============================================================
OUTPUT_PATH = 'C:/Users/lch68/Desktop/cufa-equity-report/output/삼성전자_재무데이터.xlsx'
wb.save(OUTPUT_PATH)
print(f'[OK] Excel saved: {OUTPUT_PATH}')
print(f'     시트 수: {len(wb.sheetnames)}개')
for i, name in enumerate(wb.sheetnames, 1):
    print(f'     {i:2d}. {name}')
