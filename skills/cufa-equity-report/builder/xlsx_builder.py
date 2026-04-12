"""CUFA Excel 6시트 빌더.

config 데이터 → .xlsx 파일 자동 생성.
HTML 보고서와 동일한 config 필드를 사용하므로 데이터를 2번 입력하지 않아도 된다.

6시트 구성:
    Sheet 1 — IS    손익계산서 (실적 + Forward 추정)
    Sheet 2 — BS    대차대조표
    Sheet 3 — CF    현금흐름표
    Sheet 4 — Peer  동종업체 멀티플 비교
    Sheet 5 — 추정  핵심 가정 + 시나리오 (Bull/Base/Bear)
    Sheet 6 — 밸류  Football Field 방법론 + WACC + 민감도

사용법:
    from builder.xlsx_builder import build_excel
    xlsx_path = build_excel(config)
    print(xlsx_path)  # output/HD현대중공업_CUFA_재무모델.xlsx

요구사항:
    pip install openpyxl
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ─── 디자인 상수 ────────────────────────────────────────────────────
_PURPLE   = "7C6AF7"   # CUFA 브랜드 퍼플
_DARK_BG  = "1A1A2E"   # 섹션 헤더 배경
_HEADER   = "2D2D44"   # 컬럼 헤더 배경
_POSITIVE = "4ECDC4"   # 상승/흑자 (청록)
_NEGATIVE = "FF6B6B"   # 하락/적자 (빨강)
_AMBER    = "FFD93D"   # 경고/추정치
_WHITE    = "E0E0E0"
_LIGHT    = "BBBBBB"
_SURFACE  = "0F0F0F"


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 10, color: str = _WHITE) -> "Font":
    return Font(name="맑은 고딕", bold=bold, size=size, color=color)


def _thin_border() -> "Border":
    s = Side(style="thin", color="333333")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws: Any, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_title(ws: Any, title: str, cols: int, row: int = 1) -> None:
    """시트 상단 타이틀 바."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="맑은 고딕", bold=True, size=12, color=_WHITE)
    cell.fill = _fill(_PURPLE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if cols > 1:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=cols
        )
    ws.row_dimensions[row].height = 24


def _write_header_row(ws: Any, headers: list[str], row: int) -> None:
    """컬럼 헤더 행."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _font(bold=True, size=9)
        cell.fill = _fill(_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 18


def _write_data_row(
    ws: Any,
    row: int,
    values: list,
    *,
    bold: bool = False,
    fill_hex: str | None = None,
    num_fmt: str = "#,##0",
    first_col_left: bool = True,
) -> None:
    """데이터 행 한 줄 쓰기."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = _thin_border()
        cell.font = _font(bold=bold, size=9)
        if fill_hex:
            cell.fill = _fill(fill_hex)

        if col == 1 and first_col_left:
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # 숫자 포맷
        if isinstance(val, (int, float)) and col > 1:
            cell.number_format = num_fmt


def _safe_float(v: Any) -> float | str:
    try:
        return float(v)
    except (TypeError, ValueError):
        return v if v else "-"


# ─── Sheet 1: IS ────────────────────────────────────────────────────

def _build_is_sheet(ws: Any, config: Any) -> None:
    is_cfs: list = getattr(config, "IS_CFS", []) or []
    company: str = getattr(config, "company_name", "")

    years = [r.get("year", "") for r in is_cfs]
    n_years = len(years)
    _write_title(ws, f"{company} — 손익계산서 (CFS 연결, 억원)", n_years + 1)
    _write_header_row(ws, ["구분"] + [str(y) for y in years], 2)

    ROWS = [
        ("매출액",       "revenue"),
        ("영업이익",     "operating_profit"),
        ("영업이익률(%)", "opm"),
        ("EBITDA",       "ebitda"),
        ("순이익(지배)", "net_profit"),
        ("EPS (원)",     "eps"),
        ("DPS (원)",     "dps"),
        ("ROE (%)",      "roe"),
        ("ROA (%)",      "roa"),
    ]

    for i, (label, key) in enumerate(ROWS):
        row_num = 3 + i
        is_pct = key in ("opm", "roe", "roa")
        is_rate = key in ("eps", "dps")
        fmt = "0.0%" if is_pct else ("#,##0" if not is_rate else "#,##0")
        values = [label] + [_safe_float(r.get(key, "-")) for r in is_cfs]

        # 추정 연도 강조 (E 포함)
        fill = None
        for r in is_cfs:
            if "E" in str(r.get("year", "")):
                fill = "1E1E30"  # 살짝 다른 배경
                break

        _write_data_row(ws, row_num, values, num_fmt=fmt,
                        fill_hex=fill if fill else None)

        # 영업이익 음수 → 빨강
        for col_idx, r in enumerate(is_cfs, 2):
            val = _safe_float(r.get(key, None))
            if isinstance(val, float) and val < 0:
                ws.cell(row=row_num, column=col_idx).font = Font(
                    name="맑은 고딕", size=9, color=_NEGATIVE
                )

    # 출처 행
    src_row = 3 + len(ROWS) + 1
    ws.cell(row=src_row, column=1,
            value="출처: DART 연결재무제표 / Nexus MCP. E = Forward 추정치.").font = _font(size=8, color=_LIGHT)

    _set_col_widths(ws, [18] + [12] * n_years)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "B3"


# ─── Sheet 2: BS ────────────────────────────────────────────────────

def _build_bs_sheet(ws: Any, config: Any) -> None:
    bs_cfs: list = getattr(config, "BS_CFS", []) or []
    company: str = getattr(config, "company_name", "")
    years = [r.get("year", "") for r in bs_cfs]
    n_years = max(len(years), 1)

    _write_title(ws, f"{company} — 대차대조표 (CFS 연결, 억원)", n_years + 1)
    _write_header_row(ws, ["구분"] + ([str(y) for y in years] if years else ["미입력"]), 2)

    ROWS = [
        ("── 자산 ──",            None),
        ("  유동자산",             "current_assets"),
        ("  현금및현금성자산",     "cash"),
        ("  매출채권",             "receivables"),
        ("  재고자산",             "inventory"),
        ("  비유동자산",           "non_current_assets"),
        ("  유형자산",             "ppe"),
        ("  무형자산",             "intangibles"),
        ("자산 합계",              "total_assets"),
        ("── 부채 ──",            None),
        ("  유동부채",             "current_liabilities"),
        ("  단기차입금",           "short_term_debt"),
        ("  비유동부채",           "non_current_liabilities"),
        ("  장기차입금",           "long_term_debt"),
        ("부채 합계",              "total_liabilities"),
        ("── 자본 ──",            None),
        ("  자본금",               "common_stock"),
        ("  이익잉여금",           "retained_earnings"),
        ("자본 합계 (지배)",       "total_equity"),
        ("순차입금(Net Debt)",      "net_debt"),
    ]

    for i, (label, key) in enumerate(ROWS):
        row_num = 3 + i
        if key is None:
            # 구분선 헤더 행
            ws.cell(row=row_num, column=1, value=label).font = _font(bold=True, size=9, color=_AMBER)
            ws.cell(row=row_num, column=1).fill = _fill("252535")
        else:
            values = [label] + (
                [_safe_float(r.get(key, "-")) for r in bs_cfs]
                if bs_cfs else ["-"]
            )
            bold = key in ("total_assets", "total_liabilities", "total_equity")
            _write_data_row(ws, row_num, values, bold=bold)

    src_row = 3 + len(ROWS) + 1
    ws.cell(row=src_row, column=1,
            value="출처: DART 연결재무제표 / Nexus MCP.").font = _font(size=8, color=_LIGHT)

    _set_col_widths(ws, [22] + [12] * n_years)
    ws.freeze_panes = "B3"


# ─── Sheet 3: CF ────────────────────────────────────────────────────

def _build_cf_sheet(ws: Any, config: Any) -> None:
    cf_cfs: list = getattr(config, "CF_CFS", []) or []
    company: str = getattr(config, "company_name", "")
    years = [r.get("year", "") for r in cf_cfs]
    n_years = max(len(years), 1)

    _write_title(ws, f"{company} — 현금흐름표 (CFS 연결, 억원)", n_years + 1)
    _write_header_row(ws, ["구분"] + ([str(y) for y in years] if years else ["미입력"]), 2)

    ROWS = [
        ("영업활동현금흐름 (OCF)", "ocf"),
        ("  당기순이익",            "net_income"),
        ("  감가상각",              "depreciation"),
        ("  운전자본변동",          "working_capital_change"),
        ("투자활동현금흐름 (ICF)",  "icf"),
        ("  CAPEX",                 "capex"),
        ("  투자자산변동",          "investment_change"),
        ("재무활동현금흐름 (FCF)",  "financing_cf"),
        ("잉여현금흐름 (FCF)",      "free_cf"),
        ("기말현금",                "ending_cash"),
    ]

    for i, (label, key) in enumerate(ROWS):
        row_num = 3 + i
        values = [label] + (
            [_safe_float(r.get(key, "-")) for r in cf_cfs]
            if cf_cfs else ["-"]
        )
        bold = key in ("ocf", "icf", "free_cf")
        _write_data_row(ws, row_num, values, bold=bold)

        for col_idx, r in enumerate(cf_cfs, 2):
            val = _safe_float(r.get(key, None))
            if isinstance(val, float) and val < 0 and key not in ("icf", "financing_cf", "capex", "investment_change"):
                ws.cell(row=row_num, column=col_idx).font = Font(
                    name="맑은 고딕", size=9, color=_NEGATIVE
                )

    src_row = 3 + len(ROWS) + 1
    ws.cell(row=src_row, column=1,
            value="출처: DART 연결재무제표 / Nexus MCP. FCF = OCF - CAPEX.").font = _font(size=8, color=_LIGHT)

    _set_col_widths(ws, [24] + [12] * n_years)
    ws.freeze_panes = "B3"


# ─── Sheet 4: Peer ──────────────────────────────────────────────────

def _build_peer_sheet(ws: Any, config: Any) -> None:
    peers: list = getattr(config, "PEERS", []) or []
    company: str = getattr(config, "company_name", "")
    ticker: str = getattr(config, "stock_code", "")

    _write_title(ws, f"{company} — Peer 비교 (Forward 컨센서스)", 7)
    _write_header_row(ws, ["기업", "티커", "Fwd PER", "PBR(Tr)", "ROE(%)", "OPM(%)", "비고"], 2)

    # 대상 기업 먼저
    subject_row = [
        f"★ {company}", ticker,
        _safe_float(getattr(config, "SUBJECT_FWD_PER", "-")),
        _safe_float(getattr(config, "SUBJECT_PBR", "-")),
        _safe_float(getattr(config, "SUBJECT_ROE", "-")),
        _safe_float(getattr(config, "SUBJECT_OPM", "-")),
        "분석 대상",
    ]
    _write_data_row(ws, 3, subject_row, bold=True, fill_hex=_PURPLE)

    for i, p in enumerate(peers):
        row_num = 4 + i
        values = [
            p.get("name", ""),
            p.get("ticker", p.get("code", "")),
            _safe_float(p.get("fwd_per", "-")),
            _safe_float(p.get("pbr", "-")),
            _safe_float(p.get("roe", "-")),
            _safe_float(p.get("opm", "-")),
            p.get("note", ""),
        ]
        _write_data_row(ws, row_num, values)

    # Peer 평균 행
    avg_row = 4 + len(peers)
    if peers:
        def avg(key: str) -> float | str:
            vals = [float(p[key]) for p in peers if _safe_float(p.get(key)) != "-"]
            return round(sum(vals) / len(vals), 1) if vals else "-"
        _write_data_row(
            ws, avg_row,
            ["Peer 평균", "", avg("fwd_per"), avg("pbr"), avg("roe"), avg("opm"), ""],
            bold=True, fill_hex="252535",
        )

    src_row = avg_row + 2
    ws.cell(row=src_row, column=1,
            value="출처: Bloomberg Consensus / Nexus MCP. 기준: 12MF Forward").font = _font(size=8, color=_LIGHT)

    _set_col_widths(ws, [20, 10, 10, 10, 10, 10, 20])
    ws.freeze_panes = "A4"


# ─── Sheet 5: 추정 ──────────────────────────────────────────────────

def _build_estimate_sheet(ws: Any, config: Any) -> None:
    company: str = getattr(config, "company_name", "")
    target: dict = getattr(config, "TARGET_PRICE", {}) or {}
    is_cfs: list = getattr(config, "IS_CFS", []) or []
    tt: dict = getattr(config, "trade_ticket", {}) or {}

    _write_title(ws, f"{company} — 핵심 추정 가정 & 시나리오", 5)

    # ── 핵심 가정 섹션 ──
    ws.cell(row=2, column=1, value="■ 핵심 추정 가정").font = _font(bold=True, size=10, color=_AMBER)
    _write_header_row(ws, ["가정 항목", "Bear", "Base", "Bull", "비고"], 3)

    sensitivity: list = target.get("sensitivity", []) or []
    if sensitivity:
        for i, s in enumerate(sensitivity):
            _write_data_row(ws, 4 + i, [
                s.get("variable", ""),
                s.get("bear", ""),
                s.get("base", ""),
                s.get("bull", ""),
                s.get("tp_impact", ""),
            ])
    else:
        ws.cell(row=4, column=1,
                value="TARGET_PRICE.sensitivity 리스트를 config에 입력하면 자동 채워집니다.").font = _font(size=9, color=_LIGHT)

    # ── 시나리오 TP 요약 ──
    scenario_start = max(4 + len(sensitivity), 5) + 2
    ws.cell(row=scenario_start, column=1, value="■ 시나리오별 목표주가").font = _font(bold=True, size=10, color=_AMBER)
    _write_header_row(ws, ["시나리오", "목표주가(원)", "확률", "Upside", "근거"], scenario_start + 1)

    scenarios = target.get("scenarios", []) or []
    if scenarios:
        for i, sc in enumerate(scenarios):
            _write_data_row(ws, scenario_start + 2 + i, [
                sc.get("name", ""),
                _safe_float(sc.get("price", "-")),
                sc.get("probability", ""),
                sc.get("upside", ""),
                sc.get("rationale", ""),
            ])
    else:
        # 기본 3시나리오 템플릿
        tp = _safe_float(target.get("weighted", "-"))
        entry = _safe_float(tt.get("entry_price", "-"))
        sl = _safe_float(tt.get("stop_loss", "-"))
        for i, (name, price, prob) in enumerate([
            ("Bull", tp, "25%"),
            ("Base", tp, "50%"),
            ("Bear", sl if sl != "-" else "-", "25%"),
        ]):
            upside = ""
            if isinstance(price, float) and isinstance(entry, float) and entry:
                upside = f"{(price - entry) / entry * 100:+.1f}%"
            _write_data_row(ws, scenario_start + 2 + i, [name, price, prob, upside, ""])

    _set_col_widths(ws, [22, 14, 10, 12, 30])
    ws.freeze_panes = "A4"


# ─── Sheet 6: 밸류에이션 ─────────────────────────────────────────────

def _build_valuation_sheet(ws: Any, config: Any) -> None:
    company: str = getattr(config, "company_name", "")
    target: dict = getattr(config, "TARGET_PRICE", {}) or {}
    wacc: dict = getattr(config, "WACC", {}) or {}
    tt: dict = getattr(config, "trade_ticket", {}) or {}

    _write_title(ws, f"{company} — 밸류에이션 (Football Field)", 4)

    # ── Football Field 방법론 ──
    ws.cell(row=2, column=1, value="■ 방법론별 목표주가").font = _font(bold=True, size=10, color=_AMBER)
    _write_header_row(ws, ["방법론", "산출 TP (원)", "가중치", "비고"], 3)

    methods: list = target.get("methods", []) or []
    for i, m in enumerate(methods):
        _write_data_row(ws, 4 + i, [
            m.get("method", ""),
            _safe_float(m.get("target_price", "-")),
            m.get("weight_pct", ""),
            m.get("note", ""),
        ])

    # 가중 TP 합계
    if methods:
        ws.cell(row=4 + len(methods), column=1, value="가중 목표주가").font = _font(bold=True, size=10, color=_WHITE)
        tp_cell = ws.cell(row=4 + len(methods), column=2,
                           value=_safe_float(target.get("weighted", "-")))
        tp_cell.font = Font(name="맑은 고딕", bold=True, size=12, color=_POSITIVE)
        tp_cell.number_format = "#,##0"
        ws.cell(row=4 + len(methods), column=3,
                value="100%").font = _font(bold=True, size=9)

    # ── WACC / DCF 가정 ──
    wacc_start = 4 + len(methods) + 3
    ws.cell(row=wacc_start, column=1, value="■ WACC / DCF 핵심 가정").font = _font(bold=True, size=10, color=_AMBER)
    _write_header_row(ws, ["항목", "값", "", ""], wacc_start + 1)

    wacc_items = [
        ("무위험수익률 (Rf)",       wacc.get("rf", "")),
        ("시장위험프리미엄 (ERP)",   wacc.get("erp", "")),
        ("베타 (β)",                 wacc.get("beta", "")),
        ("자기자본비용 (Ke)",        wacc.get("ke", "")),
        ("세후 타인자본비용 (Kd)",   wacc.get("kd", "")),
        ("WACC",                     wacc.get("wacc_pct", wacc.get("wacc", ""))),
        ("영구성장률 (g)",           wacc.get("terminal_growth", "")),
        ("EV/EBITDA 기준 멀티플",    wacc.get("ev_ebitda_multiple", "")),
    ]
    for i, (label, val) in enumerate(wacc_items):
        row_num = wacc_start + 2 + i
        ws.cell(row=row_num, column=1, value=label).font = _font(size=9)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=row_num, column=2, value=val).font = _font(bold=True, size=9, color=_POSITIVE)
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="left", indent=1)

    # ── Trade Ticket 요약 ──
    tt_start = wacc_start + 2 + len(wacc_items) + 2
    ws.cell(row=tt_start, column=1, value="■ Trade Ticket 요약").font = _font(bold=True, size=10, color=_AMBER)
    _write_header_row(ws, ["항목", "값", "", ""], tt_start + 1)

    tt_items = [
        ("투자의견",         tt.get("opinion", "")),
        ("진입가 (Entry)",   tt.get("entry_price", "")),
        ("목표가 (Target)",  tt.get("target_price", "")),
        ("손절가 (SL)",      tt.get("stop_loss", "")),
        ("투자 지평",        f'{tt.get("horizon_months", "")}개월' if tt.get("horizon_months") else ""),
        ("포지션 비중",      f'{tt.get("position_size_pct", "")}%' if tt.get("position_size_pct") else ""),
        ("Risk/Reward",      tt.get("risk_reward", "")),
    ]
    for i, (label, val) in enumerate(tt_items):
        row_num = tt_start + 2 + i
        ws.cell(row=row_num, column=1, value=label).font = _font(size=9)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", indent=1)
        v_cell = ws.cell(row=row_num, column=2, value=val)
        v_cell.font = _font(bold=True, size=9, color=_WHITE)
        if label == "투자의견":
            color = {"BUY": _POSITIVE, "SELL": _NEGATIVE, "HOLD": _AMBER}.get(str(val).upper(), _WHITE)
            v_cell.font = Font(name="맑은 고딕", bold=True, size=11, color=color)
        if isinstance(val, (int, float)):
            v_cell.number_format = "#,##0"
        v_cell.alignment = Alignment(horizontal="left", indent=1)

    _set_col_widths(ws, [28, 16, 12, 20])
    ws.freeze_panes = "A4"


# ─── 메인 빌더 ──────────────────────────────────────────────────────

def build_excel(config: Any, out_path: Path | str | None = None) -> Path:
    """config 데이터를 Excel 6시트 .xlsx 파일로 저장한다.

    Args:
        config:   StockConfig 인스턴스.
        out_path: 저장 경로. 없으면 `{output_dir}/{company}_CUFA_재무모델.xlsx`.

    Returns:
        생성된 .xlsx 파일 Path.

    Raises:
        ImportError: openpyxl이 설치되지 않은 경우.
    """
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl이 필요합니다: pip install openpyxl"
        )

    company: str = getattr(config, "company_name", "report")
    if out_path is None:
        out_dir = Path(getattr(config, "output_dir", "output"))
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{company}_CUFA_재무모델.xlsx"
    out_path = Path(out_path)

    wb = Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    # 시트 정의
    SHEETS = [
        ("IS",   _build_is_sheet),
        ("BS",   _build_bs_sheet),
        ("CF",   _build_cf_sheet),
        ("Peer", _build_peer_sheet),
        ("추정", _build_estimate_sheet),
        ("밸류", _build_valuation_sheet),
    ]

    for tab_name, builder_fn in SHEETS:
        ws = wb.create_sheet(title=tab_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = _PURPLE  # 탭 색상
        try:
            builder_fn(ws, config)
        except Exception as exc:
            # 개별 시트 실패 시 에러 표시 후 계속
            ws.cell(row=1, column=1, value=f"[ERROR] {exc}").font = _font(color=_NEGATIVE)

    wb.save(out_path)
    size_kb = out_path.stat().st_size // 1024
    print(f"[CUFA Excel] 저장 완료: {out_path} ({size_kb} KB, 6시트)")
    return out_path
